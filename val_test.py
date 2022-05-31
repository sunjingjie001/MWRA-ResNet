# -*- coding: utf-8 -*-
"""
Created on Wed May 25 13:51:08 2022

@author: Administrator
"""

# -*- coding: utf-8 -*-
"""
Created on Fri Apr 22 15:47:48 2022

@author: Administrator
"""
import openpyxl


import argparse
import numpy as np

from tqdm import tqdm
import tensorboard_logger
tensorboard_logger.clean_default_logger()
import torch
import torch.nn as nn
import torch.optim as optim

import torch.utils.data
from torch.utils.data import DataLoader

#tensorboard --logdir=E:\彭轶灏\EA-GAN\log

from model import ResNet
from utils import LoadDataset



def main():

	parser = argparse.ArgumentParser(description='SRGAN val')
	parser.add_argument('--device_num', default=30, type=int, help='valing date device_num')
	parser.add_argument('--Signal_num', default=500, type=int, help='valing date Signal_num')
	parser.add_argument('--start_num', default=500, type=int, help='valing date start_num')
	parser.add_argument('--val_set', default='E:/dataset/Train/dataset_training_aug_0hz.h5', type=str, help='val set path')
	parser.add_argument('--excel_set', default='E:/ResNet/291-300测试结果.xlsx', type=str, help='excel set path')
	parser.add_argument('--test_start', type=int, default=3491, help="valing date test start epoch")
	parser.add_argument('--test_num', type=int, default=10, help="valing date test epoch")
	opt = parser.parse_args()

	Signal_num=opt.Signal_num
	excel_set=opt.excel_set

	start_num=opt.start_num
	device_num=opt.device_num


	val_set = LoadDataset(opt.val_set, device_num, Signal_num, start_num)
	val_loader = DataLoader(dataset=val_set, num_workers=0, batch_size=1, shuffle=True)


	for test_epoch in range(opt.test_start,opt.test_start+opt.test_num):
		sheet_name= '第'+str(test_epoch)+'次迭代'

#打开excel文件
		book = openpyxl.load_workbook(excel_set)
		sheet = book.create_sheet(title=sheet_name,index=0)
		sheet.cell(row=1, column=1).value='实际标签'
		sheet.cell(row=1, column=2).value='预测标签'
		sheet.cell(row=1, column=3).value='准确(Flase0/True1)'
		sheet.cell(row=1, column=4).value='90置信(Flase0/True1)'





		if not torch.cuda.is_available():
			print ('!!!!!!!!!!!!!!USING CPU!!!!!!!!!!!!!')

		netResNet = ResNet(device_num)

		print('# generator parameters:', sum(param.numel() for param in netResNet.parameters()))


		#读取model
		if torch.cuda.is_available():
			netResNet.load_state_dict(torch.load('E:/ResNet/ResNet/result__2500_start/netResNet_epoch_' + str(test_epoch) + '_gpu.pth'))

		else :
			netResNet.load_state_dict(torch.load('E:/ResNet/ResNet/result__2500_start/netResNet_epoch_' + str(test_epoch) + '_gpu.pth'))
		i=1
		netResNet=netResNet.cuda()
		val_bar = tqdm(val_loader)
		for val_Signal_1, val_label in val_bar:
			if torch.cuda.is_available():
				val_Signal_1 = val_Signal_1.cuda()
				val_label=val_label.cuda()
	
	
	
			val_fake_Signal_1 = netResNet(val_Signal_1)[0, 0, :, :]
	
			i=i+1
	
	
			val_Signal_label_value=val_fake_Signal_1[0].softmax(dim=0)
			val_predict_label_probability=torch.max(val_Signal_label_value)
			val_predict_label=list(torch.where(val_Signal_label_value==torch.max(val_Signal_label_value)))[0][0]+1
			val_true_or_false= val_predict_label==val_label
			val_believe= val_predict_label_probability>=0.9
			sheet.cell(row=i, column=1).value = int(val_label)
			sheet.cell(row=i, column=2).value = int(val_predict_label)
			sheet.cell(row=i, column=3).value = int(val_true_or_false)
			sheet.cell(row=i, column=4).value = int(val_believe)
		book.save(excel_set)
#		return int(predict_success_rate)/len(val_bar)

	
	
if __name__ == '__main__':
	main()
